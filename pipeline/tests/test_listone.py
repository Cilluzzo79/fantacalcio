import pytest
from openpyxl import Workbook
from fantapipe.listone import load_listone, ListoneError


def make_xlsx(path, rows, header=("Id", "R", "Nome", "Squadra", "Qt.A", "FVM"),
              title_row=True):
    wb = Workbook()
    ws = wb.active
    if title_row:  # il file reale ha una riga titolo prima dell'header
        ws.append(["Quotazioni Fantacalcio - Stagione 2026-27"])
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def test_load_listone_normalizza(tmp_path):
    f = tmp_path / "quot.xlsx"
    make_xlsx(f, [
        [2170, "C", "Barella", "Inter", 28, 120],
        [105, "P", "Meret", "Napoli", 12, 40],
    ])
    df = load_listone(f)
    assert list(df.columns) == ["id", "nome", "ruolo", "squadra", "qta", "fvm"]
    assert len(df) == 2
    barella = df[df.id == 2170].iloc[0]
    assert barella.ruolo == "C" and barella.qta == 28


def test_header_su_prima_riga(tmp_path):
    f = tmp_path / "quot.xlsx"
    make_xlsx(f, [[1, "A", "Kean", "Fiorentina", 20, 80]], title_row=False)
    assert len(load_listone(f)) == 1


def test_ruolo_non_valido_scartato_con_warning(tmp_path):
    f = tmp_path / "quot.xlsx"
    make_xlsx(f, [
        [1, "A", "Kean", "Fiorentina", 20, 80],
        [2, "X", "Errato", "Inter", 1, 1],
    ])
    df = load_listone(f)
    assert len(df) == 1 and df.iloc[0].nome == "Kean"


def test_colonne_mancanti_errore_esplicito(tmp_path):
    f = tmp_path / "quot.xlsx"
    make_xlsx(f, [[1, "A", "Kean", "Fiorentina", 20, 80]],
              header=("Codice", "Ruolo", "Giocatore", "Team", "Prezzo", "Valore"))
    with pytest.raises(ListoneError) as e:
        load_listone(f)
    assert "Codice" in str(e.value)  # elenca le colonne trovate


def test_header_con_metadata_row_intervallata(tmp_path):
    """Header at row 3 with intervening multi-cell metadata row (content pass should find real header)"""
    f = tmp_path / "quot.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Quotazioni Fantacalcio - Stagione 2026-27"])
    ws.append(["Aggiornato al", "12/08/2026", "ore 10:00"])  # 3-cell metadata row
    ws.append(["Id", "R", "Nome", "Squadra", "Qt.A", "FVM"])  # real header at row 2
    ws.append([1, "A", "Kean", "Fiorentina", 20, 80])  # data
    wb.save(f)
    df = load_listone(f)
    assert len(df) == 1 and df.iloc[0].nome == "Kean"


def test_header_non_trovato_sparse(tmp_path):
    """Header not found: sparse data (fewer than 3 non-empty cells per row) and no id/r/nome"""
    f = tmp_path / "quot.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Titolo"])
    ws.append(["Col1"])
    ws.append(["Col2"])
    ws.append(["Value1"])
    ws.append(["Value2"])
    ws.append(["Value3"])
    wb.save(f)
    with pytest.raises(ListoneError) as e:
        load_listone(f)
    assert "Header non trovato" in str(e.value)
